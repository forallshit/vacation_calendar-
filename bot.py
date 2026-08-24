# -*- coding: utf-8 -*-
"""
Главный файл бота-напоминалки о прививках.

Как запустить:
1. pip install -r requirements.txt
2. Вставь свой токен в переменную BOT_TOKEN ниже (или задай через переменную окружения BOT_TOKEN)
3. python bot.py

Подробности — в README.md
"""

import asyncio
import io
import logging
import os
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from aiogram import Bot, Dispatcher, Router, F
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    CallbackQuery,
    FSInputFile,
    BufferedInputFile,
)

import database
import vaccines
from scheduler import setup_scheduler


# ==== Вспомогательные функции для текста (возраст, статус прививки) ====
def _plural(n: int, one: str, few: str, many: str) -> str:
    n_abs = abs(n) % 100
    n1 = n_abs % 10
    if 10 < n_abs < 20:
        return many
    if n1 == 1:
        return one
    if 2 <= n1 <= 4:
        return few
    return many


def format_duration(a, b) -> str:
    """Человеческая разница между двумя датами (порядок не важен):
    '5 дней', '3 месяца', '1 год 2 месяца' и т.д."""
    start, end = (a, b) if a <= b else (b, a)
    days = (end - start).days

    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        months -= 1
    months = max(months, 0)

    years = months // 12
    rem_months = months % 12

    if years >= 1:
        parts = [f"{years} " + _plural(years, "год", "года", "лет")]
        if rem_months > 0:
            parts.append(f"{rem_months} " + _plural(rem_months, "месяц", "месяца", "месяцев"))
        return " ".join(parts)

    if months == 0:
        return f"{days} " + _plural(days, "день", "дня", "дней")

    return f"{months} " + _plural(months, "месяц", "месяца", "месяцев")


def format_age(birth_date, today) -> str:
    """Возраст ребёнка человеческим текстом: '3 месяца', '1 год 2 месяца' и т.д."""
    return format_duration(birth_date, today)


def format_child_name(name: str) -> str:
    """Капитализирует имя ребёнка так, как его обычно пишут (Марк, Анна-Мария),
    независимо от того, как ввёл родитель (марк, МАРК, анна-мария)."""
    return name.strip().title()


def vaccine_status(v) -> tuple:
    """Возвращает (иконка, текст статуса) для прививки с учётом days_left.
    Для больших сроков (>31 дня) переводит в месяцы/года — так же, как возраст."""
    today = date.today()
    due = v["due_date"]
    days_left = v["days_left"]
    date_str = due.strftime("%d.%m.%Y")

    if days_left < 0:
        overdue = abs(days_left)
        if overdue <= 31:
            duration = f"{overdue} " + _plural(overdue, "день", "дня", "дней")
        else:
            duration = format_duration(due, today)
        return "⚪️", f"была {duration} назад ({date_str})"
    elif days_left == 0:
        return "🔴", "СЕГОДНЯ"
    elif days_left <= 31:
        icon = "🟡" if days_left <= 7 else "⚪️"
        duration = f"{days_left} " + _plural(days_left, "день", "дня", "дней")
        return icon, f"через {duration} ({date_str})"
    else:
        return "⚪️", f"через {format_duration(today, due)} ({date_str})"


def build_overdue_card(child_id: int, name: str, birth_date) -> tuple:
    """Собирает текст+клавиатуру карточки 'уже должны быть поставлены' — пересчитывается
    заново при каждой отметке, чтобы уже отмеченные прививки пропадали из списка, а
    остальные оставались на месте. Когда просроченных не осталось — карточка становится
    поздравительной с информацией о следующей предстоящей прививке."""
    today = date.today()
    name = format_child_name(name)
    schedule = vaccines.get_vaccines_for_child(birth_date, today)
    completed_ids = database.get_completed_vaccine_ids(child_id)

    overdue = [v for v in schedule if v["days_left"] < 0 and v["id"] not in completed_ids]
    overdue.sort(key=lambda v: v["days_left"])

    if overdue:
        shown = overdue[:10]
        lines = [
            f"👶 <b>{name}</b>\nВозраст: {format_age(birth_date, today)}.\n",
            f"По национальному календарю ребёнку «{name}» уже должны быть поставлены такие прививки:",
        ]
        for v in shown:
            lines.append(f"• {v['name']}")
        if len(overdue) > len(shown):
            lines.append(f"...и ещё {len(overdue) - len(shown)}")
        lines.append("\nВы уже поставили их ребёнку? Отмечайте по одной ✅")
        text = "\n".join(lines)

        buttons = [
            [InlineKeyboardButton(text=f"✅ {v['name']}", callback_data=f"overduedone:{child_id}:{v['id']}")]
            for v in shown
        ]
        buttons.append([InlineKeyboardButton(text="📋 Предстоящие прививки", callback_data=f"next6:{child_id}")])
        return text, InlineKeyboardMarkup(inline_keyboard=buttons)

    # просроченных не осталось — поздравляем и показываем следующую предстоящую
    upcoming = [v for v in schedule if v["id"] not in completed_ids]
    upcoming.sort(key=lambda v: v["days_left"])

    if upcoming:
        nxt = upcoming[0]
        icon, status = vaccine_status(nxt)
        text = (
            f"🎉 Отлично! У ребёнка «{name}» стоят все нужные прививки по возрасту.\n\n"
            f"Следующая: <b>{nxt['name']}</b> — {icon} {status}\n"
            f"Я напомню о ней заранее, за неделю."
        )
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="ℹ️ Подробнее об этой прививке", callback_data=f"vaccinedetail:{nxt['id']}:{child_id}")],
            [InlineKeyboardButton(text="Спасибо! ❣️", callback_data="thanks_ack")],
        ])
    else:
        text = f"🎉 Отлично! У ребёнка «{name}» поставлены вообще все прививки по национальному календарю."
        keyboard = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="Спасибо! ❣️", callback_data="thanks_ack")
        ]])

    return text, keyboard

# ==== НАСТРОЙКИ ====
# Вставь сюда токен, который дал @BotFather, ИЛИ задай переменную окружения BOT_TOKEN
BOT_TOKEN = os.getenv("BOT_TOKEN", "ВСТАВЬ_СЮДА_СВОЙ_ТОКЕН")

# Картинка с инфографикой календаря прививок, которая приходит по /start
ASSETS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
CALENDAR_IMAGE_PATH = os.path.join(ASSETS_DIR, "vaccine_calendar.png")

# Название партнёра (роддома/клиники), которое подставляется в приветствие.
# Настраивается через переменную окружения ORG_NAME в Railway — код трогать не нужно.
# Если переменная не задана, строка с названием партнёра просто не показывается.
ORG_NAME = os.getenv("ORG_NAME", "")

# Telegram ID сотрудников клиники, у которых есть доступ к админ-панели (/admin).
# В Railway задаётся через переменную ADMIN_IDS — несколько ID через запятую,
# например: 123456789,987654321. Свой ID можно узнать командой /whoami.
ADMIN_IDS = {
    int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()
}

logging.basicConfig(level=logging.INFO)

router = Router()


# ==== Постоянное меню с кнопками внизу экрана ====
main_menu = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="👶 Мои дети")],
        [KeyboardButton(text="➕ Добавить ребёнка"), KeyboardButton(text="✅ Отметить прививку")],
    ],
    resize_keyboard=True,
)


# ==== Состояния диалога добавления ребёнка (FSM — машина состояний) ====
class AddChild(StatesGroup):
    waiting_for_name = State()
    waiting_for_birth_date = State()


# ==== Команда /start ====
@router.message(CommandStart())
async def cmd_start(message: Message):
    # при каждом /start список детей этого пользователя обнуляется с нуля
    database.delete_all_children(message.from_user.id)

    caption = (
        "Этот бот поможет вам не пропустить обязательные прививки по "
        "национальному календарю РФ.\n\n"
        "Добавьте ребёнка — и бот покажет, какие прививки пора ставить, а какие "
        "ещё предстоят, и сам напомнит заранее, когда придёт время."
    )
    if ORG_NAME:
        caption += f"\n\nСервис предоставлен: {ORG_NAME}"
    add_button = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="➕ Добавить ребёнка", callback_data="start_add_child")
    ]])

    photo_sent = False
    if os.path.isfile(CALENDAR_IMAGE_PATH):
        try:
            photo = FSInputFile(CALENDAR_IMAGE_PATH)
            await message.answer_photo(photo=photo, caption=caption, reply_markup=add_button)
            photo_sent = True
        except Exception as e:
            logging.warning(f"Не удалось отправить картинку календаря: {e}")

    if not photo_sent:
        # на случай, если картинка почему-то не попала в деплой — бот не должен падать
        await message.answer(caption, reply_markup=add_button)

    await message.answer(
        "Кнопки ниже всегда под рукой — чтобы посмотреть детей, добавить "
        "нового или отметить прививку сделанной:",
        reply_markup=main_menu,
    )


@router.message(Command("help"))
async def cmd_help(message: Message):
    await message.answer(
        "Я слежу за графиком прививок по датам рождения детей, которых вы добавите.\n"
        "Раз в день проверяю, у кого скоро прививка, и присылаю напоминание заранее.\n\n"
        "👶 Мои дети — выберите ребёнка, дальше увидите его предстоящие и уже "
        "поставленные прививки. У каждой прививки — описание, от чего она и "
        "почему важна.\n"
        "➕ Добавить ребёнка\n"
        "✅ Отметить прививку сделанной (чтобы не напоминал зря)\n"
        "/list — весь национальный календарь прививок справочником",
        reply_markup=main_menu,
    )


# ==== Справочник: весь национальный календарь прививок (не привязан к ребёнку) ====
def build_vaccine_list_keyboard():
    buttons = [
        [InlineKeyboardButton(text=v["name"], callback_data=f"scheduleinfo:{v['id']}")]
        for v in vaccines.get_all_vaccines()
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


@router.message(Command("list"))
async def cmd_list_all(message: Message):
    await message.answer(
        "💉 Полный национальный календарь прививок.\nВыберите прививку, чтобы узнать, "
        "на каком сроке она ставится, от чего защищает и почему важна:",
        reply_markup=build_vaccine_list_keyboard(),
    )


@router.callback_query(F.data.startswith("scheduleinfo:"))
async def on_schedule_info(callback: CallbackQuery):
    vaccine_id = callback.data.split(":", 1)[1]
    vaccine = vaccines.get_vaccine_by_id(vaccine_id)

    if vaccine is None:
        await callback.answer("Не нашёл информацию об этой прививке", show_alert=True)
        return

    text = (
        f"💉 <b>{vaccine['name']}</b>\n"
        f"📅 Когда ставится: {vaccine.get('age_label', '—')}\n\n"
        f"{vaccine['info']}"
    )
    keyboard = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="⬅️ Ко всему календарю", callback_data="list_all_back")
    ]])
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()


@router.callback_query(F.data == "list_all_back")
async def on_list_all_back(callback: CallbackQuery):
    await callback.message.edit_text(
        "💉 Полный национальный календарь прививок.\nВыберите прививку, чтобы узнать, "
        "на каком сроке она ставится, от чего защищает и почему важна:",
        reply_markup=build_vaccine_list_keyboard(),
    )
    await callback.answer()


# ==== Добавление ребёнка: шаг 1 — имя ====
async def ask_child_name(send_func, state: FSMContext):
    """Общий код для запроса имени ребёнка — используется и из команды/кнопки
    меню, и из кнопки под приветственной картинкой."""
    await send_func("Как зовут малыша?", reply_markup=ReplyKeyboardRemove())
    await state.set_state(AddChild.waiting_for_name)


@router.message(Command("add"))
@router.message(F.text == "➕ Добавить ребёнка")
async def cmd_add(message: Message, state: FSMContext):
    await ask_child_name(message.answer, state)


@router.callback_query(F.data == "start_add_child")
async def on_start_add_child(callback: CallbackQuery, state: FSMContext):
    await ask_child_name(callback.message.answer, state)
    await callback.answer()


@router.message(AddChild.waiting_for_name)
async def process_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text.strip())
    await message.answer(
        "Дата рождения в формате ДД.ММ.ГГГГ (например, 15.03.2024):"
    )
    await state.set_state(AddChild.waiting_for_birth_date)


# ==== Добавление ребёнка: шаг 2 — дата рождения ====
@router.message(AddChild.waiting_for_birth_date)
async def process_birth_date(message: Message, state: FSMContext):
    raw = message.text.strip()
    try:
        birth_date = datetime.strptime(raw, "%d.%m.%Y").date()
    except ValueError:
        await message.answer(
            "Не получилось распознать дату. Введите в формате ДД.ММ.ГГГГ, "
            "например: 15.03.2024"
        )
        return

    if birth_date > date.today():
        await message.answer("Дата рождения не может быть в будущем. Попробуйте ещё раз.")
        return

    data = await state.get_data()
    name = format_child_name(data["name"])

    database.add_child(
        telegram_user_id=message.from_user.id,
        name=name,
        birth_date=birth_date.isoformat(),
        parent_name=message.from_user.full_name,
        parent_username=message.from_user.username,
    )
    child_id = None
    # находим только что добавленного ребёнка (совпадение по имени и дате рождения)
    for cid, n, b in database.get_children(message.from_user.id):
        if n == name and b == birth_date.isoformat():
            child_id = cid

    await state.clear()

    today = date.today()
    schedule = vaccines.get_vaccines_for_child(birth_date, today)
    has_overdue = any(v["days_left"] < 0 for v in schedule)

    if has_overdue:
        # ребёнку больше нескольких дней — по графику уже должны быть прививки,
        # которые могли быть поставлены раньше (например, ещё в роддоме)
        text, keyboard = build_overdue_card(child_id, name, birth_date)
    else:
        upcoming = [v for v in schedule if v["days_left"] >= 0]
        upcoming.sort(key=lambda v: v["days_left"])
        next_vaccine = upcoming[0]
        icon, status = vaccine_status(next_vaccine)
        text = (
            f"👶 Привет, {name}!\nТебе сегодня: {format_age(birth_date, today)}.\n\n"
            f"💉 Ближайшая прививка: <b>{next_vaccine['name']}</b>\n"
            f"{icon} {status}"
        )
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(
                text="✅ Поставили",
                callback_data=f"markdone:{child_id}:{next_vaccine['id']}",
            )],
            [InlineKeyboardButton(text="📋 Посмотреть следующие", callback_data=f"next6:{child_id}")],
        ])

    await message.answer(text, reply_markup=keyboard)


# ==== Вспомогательное: клавиатура со списком детей ====
def build_children_keyboard(children):
    buttons = [
        [InlineKeyboardButton(text=f"👶 {format_child_name(name)}", callback_data=f"childmenu:{child_id}")]
        for child_id, name, _ in children
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


# ==== Список детей — тап открывает меню ребёнка ====
@router.message(Command("children"))
@router.message(Command("kids"))
@router.message(F.text == "👶 Мои дети")
async def cmd_children(message: Message):
    children = database.get_children(message.from_user.id)

    if not children:
        await message.answer(
            "У вас пока нет добавленных детей. Добавьте через «➕ Добавить ребёнка»",
            reply_markup=main_menu,
        )
        return

    await message.answer("Выберите ребёнка:", reply_markup=build_children_keyboard(children))


@router.callback_query(F.data == "backchildren")
async def on_back_children(callback: CallbackQuery):
    children = database.get_children(callback.from_user.id)

    if not children:
        await callback.message.edit_text(
            "У вас пока нет добавленных детей. Добавьте через «➕ Добавить ребёнка»"
        )
        await callback.answer()
        return

    await callback.message.edit_text("Выберите ребёнка:", reply_markup=build_children_keyboard(children))
    await callback.answer()


# ==== Карточка ребёнка: поставленные + ближайшие 3 предстоящие ====
@router.callback_query(F.data.startswith("childmenu:"))
async def on_child_menu(callback: CallbackQuery):
    child_id = int(callback.data.split(":")[1])
    children = database.get_children(callback.from_user.id)
    match = next(((n, b) for cid, n, b in children if cid == child_id), None)

    if match is None:
        await callback.answer("Не нашёл такого ребёнка", show_alert=True)
        return

    name, birth_date_str = match
    name = format_child_name(name)
    birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d").date()
    today = date.today()

    schedule = vaccines.get_vaccines_for_child(birth_date, today)
    completed_ids = database.get_completed_vaccine_ids(child_id)
    completed = database.get_completed_vaccines_with_dates(child_id)

    upcoming = [v for v in schedule if v["id"] not in completed_ids]
    upcoming.sort(key=lambda v: v["days_left"])
    next3 = upcoming[:3]

    lines = [f"👶 <b>{name}</b>\nВозраст: {format_age(birth_date, today)}\n"]

    if completed:
        lines.append("✅ <b>Поставленные прививки:</b>")
        for vaccine_id, completed_date in completed:
            vaccine_name = vaccines.get_vaccine_name_by_id(vaccine_id)
            date_formatted = datetime.strptime(completed_date, "%Y-%m-%d").strftime("%d.%m.%Y")
            lines.append(f"• {vaccine_name} — {date_formatted}")
    else:
        lines.append("✅ Пока нет отмеченных прививок.")

    lines.append("")

    if next3:
        lines.append("💉 <b>Ближайшие предстоящие:</b>")
        for v in next3:
            icon, status = vaccine_status(v)
            lines.append(f"{icon} {v['name']} — {status}")
    else:
        lines.append("💉 Все прививки по графику уже поставлены. 🎉")

    text = "\n".join(lines)

    buttons = [
        [InlineKeyboardButton(text=v["name"], callback_data=f"vaccinedetail:{v['id']}:{child_id}")]
        for v in next3
    ]
    if len(upcoming) > len(next3):
        buttons.append([InlineKeyboardButton(text="📋 Все предстоящие", callback_data=f"upcoming:{child_id}")])
    buttons.append([InlineKeyboardButton(text="⬅️ К списку детей", callback_data="backchildren")])

    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()


# ==== Предстоящие прививки конкретного ребёнка (список кнопками) ====
@router.callback_query(F.data.startswith("upcoming:"))
async def on_upcoming(callback: CallbackQuery):
    child_id = int(callback.data.split(":")[1])
    children = database.get_children(callback.from_user.id)
    match = next(((n, b) for cid, n, b in children if cid == child_id), None)

    if match is None:
        await callback.answer("Не нашёл такого ребёнка", show_alert=True)
        return

    name, birth_date_str = match
    name = format_child_name(name)
    birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d").date()
    schedule = vaccines.get_vaccines_for_child(birth_date, date.today())
    completed_ids = database.get_completed_vaccine_ids(child_id)

    not_done = [v for v in schedule if v["id"] not in completed_ids]
    not_done.sort(key=lambda v: v["days_left"])

    back_button = InlineKeyboardButton(text="⬅️ Назад", callback_data=f"childmenu:{child_id}")

    if not not_done:
        keyboard = InlineKeyboardMarkup(inline_keyboard=[[back_button]])
        await callback.message.edit_text(
            f"У ребёнка «{name}» все прививки по графику уже поставлены. 🎉", reply_markup=keyboard
        )
        await callback.answer()
        return

    buttons = []
    for v in not_done[:10]:
        icon, _ = vaccine_status(v)
        buttons.append([InlineKeyboardButton(
            text=f"{icon} {v['name']}",
            callback_data=f"vaccinedetail:{v['id']}:{child_id}",
        )])
    buttons.append([back_button])

    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    await callback.message.edit_text(
        f"💉 <b>{name}</b> — предстоящие прививки:\nВыберите прививку, чтобы узнать подробности.",
        reply_markup=keyboard,
    )
    await callback.answer()


# ==== Прививки на ближайшие полгода (после добавления ребёнка) ====
@router.callback_query(F.data.startswith("next6:"))
async def on_next6(callback: CallbackQuery):
    child_id = int(callback.data.split(":")[1])
    children = database.get_children(callback.from_user.id)
    match = next(((n, b) for cid, n, b in children if cid == child_id), None)

    if match is None:
        await callback.answer("Не нашёл такого ребёнка", show_alert=True)
        return

    name, birth_date_str = match
    name = format_child_name(name)
    birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d").date()
    schedule = vaccines.get_vaccines_for_child(birth_date, date.today())
    completed_ids = database.get_completed_vaccine_ids(child_id)

    upcoming6 = [v for v in schedule if v["id"] not in completed_ids and v["days_left"] <= 183]
    upcoming6.sort(key=lambda v: v["days_left"])

    back_button = InlineKeyboardButton(text="⬅️ Назад", callback_data=f"childmenu:{child_id}")

    if not upcoming6:
        keyboard = InlineKeyboardMarkup(inline_keyboard=[[back_button]])
        await callback.message.edit_text(
            f"У ребёнка «{name}» нет прививок по графику в ближайшие полгода.", reply_markup=keyboard
        )
        await callback.answer()
        return

    buttons = []
    for v in upcoming6:
        icon, _ = vaccine_status(v)
        buttons.append([InlineKeyboardButton(
            text=f"{icon} {v['name']}",
            callback_data=f"vaccinedetail:{v['id']}:{child_id}",
        )])
    buttons.append([back_button])

    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    await callback.message.edit_text(
        f"💉 <b>{name}</b> — прививки на ближайшие полгода:", reply_markup=keyboard
    )
    await callback.answer()


# ==== Карточка одной прививки: описание + кнопка "Поставили" ====
@router.callback_query(F.data.startswith("vaccinedetail:"))
async def on_vaccine_detail(callback: CallbackQuery):
    _, vaccine_id, child_id_str = callback.data.split(":")
    child_id = int(child_id_str)
    vaccine = vaccines.get_vaccine_by_id(vaccine_id)

    if vaccine is None:
        await callback.answer("Не нашёл информацию об этой прививке", show_alert=True)
        return

    text = f"💉 <b>{vaccine['name']}</b>\n\n{vaccine['info']}"
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Поставили", callback_data=f"markdone:{child_id}:{vaccine_id}")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data=f"upcoming:{child_id}")],
    ])
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()


# ==== Поставленные прививки конкретного ребёнка ====
@router.callback_query(F.data.startswith("donelist:"))
async def on_done_list(callback: CallbackQuery):
    child_id = int(callback.data.split(":")[1])
    children = database.get_children(callback.from_user.id)
    name = next((n for cid, n, _ in children if cid == child_id), None)

    if name is None:
        await callback.answer("Не нашёл такого ребёнка", show_alert=True)
        return

    name = format_child_name(name)
    back_button = InlineKeyboardButton(text="⬅️ Назад", callback_data=f"childmenu:{child_id}")
    keyboard = InlineKeyboardMarkup(inline_keyboard=[[back_button]])

    completed = database.get_completed_vaccines_with_dates(child_id)
    if not completed:
        await callback.message.edit_text(
            f"У ребёнка «{name}» пока нет поставленных прививок.", reply_markup=keyboard
        )
        await callback.answer()
        return

    lines = [f"✅ <b>{name}</b> — поставленные прививки:\n"]
    for vaccine_id, completed_date in completed:
        vaccine_name = vaccines.get_vaccine_name_by_id(vaccine_id)
        date_formatted = datetime.strptime(completed_date, "%Y-%m-%d").strftime("%d.%m.%Y")
        lines.append(f"• {vaccine_name} — {date_formatted}")

    await callback.message.edit_text("\n".join(lines), reply_markup=keyboard)
    await callback.answer()


# ==== Отметить прививку сделанной — быстрый способ через кнопку меню ====
@router.message(Command("done"))
@router.message(F.text == "✅ Отметить прививку")
async def cmd_done(message: Message):
    children = database.get_children(message.from_user.id)

    if not children:
        await message.answer(
            "У вас пока нет добавленных детей. Добавьте через «➕ Добавить ребёнка»",
            reply_markup=main_menu,
        )
        return

    if len(children) == 1:
        child_id, name, _ = children[0]
        keyboard = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="💉 Предстоящие прививки", callback_data=f"upcoming:{child_id}")
        ]])
        await message.answer(f"👶 <b>{format_child_name(name)}</b>", reply_markup=keyboard)
    else:
        await message.answer("Какого ребёнка?", reply_markup=build_children_keyboard(children))


# ==== Отметить прививку сделанной (сама отметка) ====
@router.callback_query(F.data.startswith("markdone:"))
async def on_mark_done(callback: CallbackQuery):
    _, child_id_str, vaccine_id = callback.data.split(":")
    child_id = int(child_id_str)

    database.mark_vaccine_done(
        child_id=child_id,
        vaccine_id=vaccine_id,
        completed_date=date.today().isoformat(),
    )

    vaccine_name = vaccines.get_vaccine_name_by_id(vaccine_id)

    children = database.get_children(callback.from_user.id)
    match = next(((n, b) for cid, n, b in children if cid == child_id), None)

    next_line = ""
    child_name = ""
    if match:
        child_name, birth_date_str = match
        child_name = format_child_name(child_name)
        birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d").date()
        schedule = vaccines.get_vaccines_for_child(birth_date, date.today())
        completed_ids = database.get_completed_vaccine_ids(child_id)
        not_done = [v for v in schedule if v["id"] not in completed_ids]

        if not_done:
            not_done.sort(key=lambda v: v["days_left"])
            nxt = not_done[0]
            _, status = vaccine_status(nxt)
            next_line = f"\n\nСледующая прививка запланирована: <b>{nxt['name']}</b> — {status}"
        else:
            next_line = "\n\nБольше прививок по графику нет — все поставлены. 🎉"

    text = f"✅ Отлично! Теперь {child_name} защищён(а) от: <b>{vaccine_name}</b>.{next_line}"
    keyboard = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="⬅️ К прививкам ребёнка", callback_data=f"childmenu:{child_id}")
    ]])
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer("Готово!")


# ==== Отметить одну прививку из списка "уже должны быть поставлены" ====
# В отличие от markdone — не заменяет карточку целиком, а пересчитывает список:
# отмеченная прививка пропадает, остальные кнопки остаются на месте.
@router.callback_query(F.data.startswith("overduedone:"))
async def on_overdue_done(callback: CallbackQuery):
    _, child_id_str, vaccine_id = callback.data.split(":")
    child_id = int(child_id_str)

    database.mark_vaccine_done(
        child_id=child_id,
        vaccine_id=vaccine_id,
        completed_date=date.today().isoformat(),
    )

    children = database.get_children(callback.from_user.id)
    match = next(((n, b) for cid, n, b in children if cid == child_id), None)

    if match is None:
        await callback.answer("Не нашёл такого ребёнка", show_alert=True)
        return

    name, birth_date_str = match
    birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d").date()

    text, keyboard = build_overdue_card(child_id, name, birth_date)
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer("Отмечено ✅")


@router.callback_query(F.data == "thanks_ack")
async def on_thanks_ack(callback: CallbackQuery):
    await callback.answer("Пожалуйста! 🍀")


# ==== Узнать свой Telegram ID (нужно, чтобы настроить ADMIN_IDS) ====
@router.message(Command("whoami"))
async def cmd_whoami(message: Message):
    await message.answer(f"Ваш Telegram ID: <code>{message.from_user.id}</code>")


# ==== Админ-панель для персонала клиники: список всех пациентов ====
@router.message(Command("admin"))
async def cmd_admin(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("Эта команда доступна только персоналу клиники.")
        return

    children = database.get_all_children_full()
    if not children:
        await message.answer("Пока нет ни одного добавленного ребёнка.")
        return

    lines = [f"📋 <b>Всего пациентов: {len(children)}</b>\n"]
    for child_id, name, birth_date_str, parent_name, parent_username in children:
        birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d").date()
        completed = database.get_completed_vaccines_with_dates(child_id)

        parent_line = parent_name or "—"
        if parent_username:
            parent_line += f" (@{parent_username})"

        lines.append(f"👩 {parent_line}")
        lines.append(f"👶 {format_child_name(name)} — {birth_date.strftime('%d.%m.%Y')}")
        if completed:
            for vaccine_id, completed_date in completed:
                vaccine_name = vaccines.get_vaccine_name_by_id(vaccine_id)
                date_formatted = datetime.strptime(completed_date, "%Y-%m-%d").strftime("%d.%m.%Y")
                lines.append(f"   ✅ {vaccine_name} — {date_formatted}")
        else:
            lines.append("   — прививок пока не отмечено")
        lines.append("")

    # Разбиваем на несколько сообщений, чтобы не упереться в лимит Telegram (~4096 символов)
    chunk, chunk_len = [], 0
    for line in lines:
        if chunk_len + len(line) + 1 > 3500:
            await message.answer("\n".join(chunk))
            chunk, chunk_len = [], 0
        chunk.append(line)
        chunk_len += len(line) + 1
    if chunk:
        await message.answer("\n".join(chunk))


# ==== Выгрузка списка пациентов файлом Excel ====
@router.message(Command("admin_export"))
async def cmd_admin_export(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("Эта команда доступна только персоналу клиники.")
        return

    children = database.get_all_children_full()
    if not children:
        await message.answer("Пока нет ни одного добавленного ребёнка.")
        return

    today = date.today()
    NAME_FILL = PatternFill(start_color="1466AF", end_color="1466AF", fill_type="solid")
    NAME_FONT = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    SUB_FONT = Font(name="Arial", size=10, italic=True, color="FFFFFF")
    DONE_HEAD_FILL = PatternFill(start_color="D9F2E1", end_color="D9F2E1", fill_type="solid")
    NEXT_HEAD_FILL = PatternFill(start_color="FDEEF0", end_color="FDEEF0", fill_type="solid")
    COL_HEAD_FONT = Font(name="Arial", size=11, bold=True, color="10233B")
    BODY_FONT = Font(name="Arial", size=10)

    wb = Workbook()
    ws = wb.active
    ws.title = "Пациенты"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42

    row = 1
    for child_id, name, birth_date_str, parent_name, parent_username in children:
        birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d").date()
        display_name = format_child_name(name)
        username_fmt = f"@{parent_username}" if parent_username else "—"
        parent_display = parent_name or "—"

        completed = database.get_completed_vaccines_with_dates(child_id)
        schedule = vaccines.get_vaccines_for_child(birth_date, today)
        completed_ids = database.get_completed_vaccine_ids(child_id)
        upcoming = [v for v in schedule if v["id"] not in completed_ids]
        upcoming.sort(key=lambda v: v["days_left"])

        # ==== Заголовок блока: имя ребёнка + мама ====
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=f"👶 {display_name}   🤱 {parent_display} ({username_fmt})")
        cell.font = NAME_FONT
        cell.fill = NAME_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 26
        row += 1

        # ==== Подстрока: дата рождения + возраст ====
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(
            row=row, column=1,
            value=f"Дата рождения: {birth_date.strftime('%d.%m.%Y')}   •   Возраст: {format_age(birth_date, today)}",
        )
        cell.font = SUB_FONT
        cell.fill = NAME_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        # ==== Заголовки двух колонок ====
        c1 = ws.cell(row=row, column=1, value="✅ Поставлено")
        c1.font = COL_HEAD_FONT
        c1.fill = DONE_HEAD_FILL
        c1.alignment = Alignment(horizontal="left", vertical="center")

        c2 = ws.cell(row=row, column=2, value="📋 Предстоит")
        c2.font = COL_HEAD_FONT
        c2.fill = NEXT_HEAD_FILL
        c2.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        completed_lines = []
        for vaccine_id, completed_date in completed:
            vaccine_name = vaccines.get_vaccine_name_by_id(vaccine_id)
            date_fmt = datetime.strptime(completed_date, "%Y-%m-%d").strftime("%d.%m.%Y")
            completed_lines.append(f"{vaccine_name} — {date_fmt}")

        upcoming_lines = []
        for v in upcoming:
            _, status = vaccine_status(v)
            upcoming_lines.append(f"{v['name']} — {status}")

        if not completed_lines:
            completed_lines = ["—"]
        if not upcoming_lines:
            upcoming_lines = ["все прививки поставлены 🎉"]

        max_len = max(len(completed_lines), len(upcoming_lines))
        for i in range(max_len):
            left = completed_lines[i] if i < len(completed_lines) else ""
            right = upcoming_lines[i] if i < len(upcoming_lines) else ""
            ws.cell(row=row, column=1, value=left).font = BODY_FONT
            ws.cell(row=row, column=2, value=right).font = BODY_FONT
            row += 1

        row += 1  # пустая строка-разделитель между детьми

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    file = BufferedInputFile(buffer.read(), filename="patients.xlsx")
    await message.answer_document(file, caption=f"📊 Выгрузка пациентов: {len(children)} детей")


async def main():
    database.init_db()

    bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)

    # Планировщик, который раз в день шлёт напоминания (см. scheduler.py)
    setup_scheduler(bot)

    logging.info("Бот запущен")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
